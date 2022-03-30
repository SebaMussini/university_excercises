#---------------------************* FUENTES ***********--------------------


# https://www.geeksforgeeks.org/python-reading-excel-file-using-openpyxl-module/

# https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f

# https://stackoverflow.com/questions/41556378/openpyxl-set-active-sheet/50117733

# https://relopezbriega.github.io/blog/2016/09/26/series-de-tiempo-con-python/

# https://zetcode.com/python/openpyxl/

# https://stackoverflow.com/questions/29354868/how-can-openpyxl-write-list-data-in-sheet


#------------------------------*********************------------------------

# Para leer un archivo de excel (xlsx), utilizaremos la librería
# openpyxl


# Importamos la librería

import openpyxl

from openpyxl import Workbook

from openpyxl import load_workbook

from statistics import mean

import pandas as pd

import matplotlib.pyplot as plt


# Lo siguiente que haremos será indicarle al programa el lugar donde se
# encuentra el archivo de excel con el cual vamos a trabajar.

path = "/home/seba/Escritorio/Clases UM 2022/Python_SQL_Clases/Macro I/Actividades_IVF.xlsx"


# Una vez realizado lo anterior, debemos crear un objeto en el cual
# que abra lo que existe en el path. A este objeto, lo llamaremos
#   "wb_obj"  // "WorkBook"

wb_obj = openpyxl.load_workbook(path)

# Quiero ver todas las solapas que tiene el achivo excel

print(wb_obj.sheetnames)


# La hoja/solapa activa es aquella en la que estaba posicionado al guardar el
# archivo excel

print(wb_obj.active)




# Ahora selecciono la solapa "IVF" para trabajar sobre esa informaciónote

ws_IVF = wb_obj['IVF']
print(ws_IVF)


# Creamos dos listas que utilizaremos a continuación, para crear una serie
# con los trimestres y sus correspondientes valores

serie_IVF_nombre = list()
serie_IVF_valor = list()


# Definimos una variable que represente el numero máximo de columnas que
# existen en el archivo de excel.

max_col = ws_IVF.max_column

# Creamos una variable Count  y   Sum  que comiencen con valor cero

count = 0
suma = 0


# Iteramos de forma tal que para cada una de las "i" columnas, busque el valor
# que coincidan con la fila 8 (la cual contiene los Trimestre-Año), y lo
# agregue a la lista con nombre "serie_IVF_nombre"

for i in range(1, max_col + 1):
    cell_obj = ws_IVF.cell(row = 8, column = i)
    suma += count + 1
    #print(cell_obj.value)
    #print(suma)
    serie_IVF_nombre.append(cell_obj.value)


# Idem el bucle anterior pero ahora iterando con la fila que contiene NO los
# Trimestre-Año, sino que contiene los valores del IVF que corresponde.

for i in range(1, max_col + 1):
    cell_obj = ws_IVF.cell(row = 21, column = i)
    suma += count + 1
    #print(cell_obj.value)
    #print(suma)
    serie_IVF_valor.append(cell_obj.value)

# print(serie_IVF)


# Ahora le agregamos a la serie un valor específico en el index/posición 5,
# con el título de la serie.

serie_IVF_nombre.insert(5, "Trimestre-Año")
serie_IVF_valor.insert(5, "IVF del PIB")


# Ahora vamos a re-definir dos series como sub-conjuntos de las anteriores, de
# manera tal que nos muestre una parte de la información (desde el index = 5
# al final) y no toda.

serie_IVF_PIB_nombre = serie_IVF_nombre[6:]
serie_IVF_PIB_valor = serie_IVF_valor[6:]
print("La serie del IVF del PIB es:", serie_IVF_PIB_valor)






def fpp_3():



    # Observar que el comando es el mismo que en el primer ejemplo,
    # "plt.plot", solamente que ahora estamos agregando más argumentos
    # al comando (colores de línea, tipo de línea, etc.)
    plt.plot(serie_IVF_PIB_nombre, serie_IVF_PIB_valor, color='red', linestyle='dashdot', linewidth = 3,
             marker='*', markerfacecolor='yellow', markersize=12)



    # Podemos elegir el rango de cada eje, para visualizar mejor el gráfico

    plt.xlabel('Trimestre-Año')

    plt.ylabel('IVF')

    plt.title('Evolución del IVF del PIB base 2016')

    plt.show()

fpp_3()
