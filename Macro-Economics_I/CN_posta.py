#---------------------************* FUENTES ***********--------------------


# https://www.geeksforgeeks.org/python-reading-excel-file-using-openpyxl-module/

# https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f

# https://stackoverflow.com/questions/41556378/openpyxl-set-active-sheet/50117733

# https://relopezbriega.github.io/blog/2016/09/26/series-de-tiempo-con-python/

# https://zetcode.com/python/openpyxl/

# https://stackoverflow.com/questions/29354868/how-can-openpyxl-write-list-data-in-sheet

# https://www.delftstack.com/es/howto/python/moving-average-python/


#------------------------------*********************------------------------

# Para leer un archivo de excel (xlsx), utilizaremos la librería
# openpyxl


# Importamos la librería

import openpyxl

from openpyxl import Workbook

from openpyxl import load_workbook

from statistics import mean

import pandas as pd

import matplotlib.pyplot as plt

import numpy as np



# Lo siguiente que haremos será indicarle al programa el lugar donde se
# encuentra el archivo de excel con el cual vamos a trabajar.

path = "/home/seba/Escritorio/Clases UM 2022/Python_SQL_Clases/Macro I/Actividades_IVF.xlsx"


# Una vez realizado lo anterior, debemos crear un objeto en el cual
# que abra lo que existe en el path. A este objeto, lo llamaremos
#   "wb_obj"  // "WorkBook"

wb_obj = openpyxl.load_workbook(path)

# Quiero ver todas las solapas que tiene el achivo excel

print(wb_obj.sheetnames)


# La hoja/solapa activa es aquella en la que estaba posicionado al guardar el
# archivo excel

print(wb_obj.active)

# Creo una nueva hoja/solapa con el nombre de "Seba", posicionada como
# tercera solapa

ws_obj = wb_obj.create_sheet('Seba', 3)  # insert at index 3
print(wb_obj.sheetnames)
wb_obj.save('Actividades_IVF.xlsx')


# Selecciono / me posiciono en cada una de las diferentes solapas.

# Puedo llamarlas por su index en el woorkbook (wb)  o bien puedo crear
# un objeto woorksheet (ws) que sea la solapa que deseo

wb_obj.active = 0
print(wb_obj.active)

ws_IVF = wb_obj['IVF']
print(ws_IVF)




wb_obj.active = 1
print(wb_obj.active)

ws_IVF_anual = wb_obj['IVF_anual']
print(ws_IVF_anual)




wb_obj.active = 2
print(wb_obj.active)

ws_Seba = wb_obj['Seba']
#print(ws_Seba)




wb_obj.active = 3
print(wb_obj.active)


# Ahora selecciono la solapa "IVF" para trabajar sobre esa informaciónote

ws_IVF = wb_obj['IVF']
print(ws_IVF)


# Mostrar dimensiones de la matriz, filas y columnas

print("Cantidad total de  filas: ", ws_IVF.max_row)

print("Cantidad total de columnas: ", ws_IVF.max_column)


# Observamos un valor particular

cell_obj = ws_IVF.cell(row = 21, column = 28)

#print(cell_obj.value)



# Creamos dos listas que utilizaremos a continuación, para crear una serie
# con los trimestres y sus correspondientes valores

serie_IVF_nombre = list()
serie_IVF_valor = list()


# Definimos una variable que represente el numero máximo de columnas que
# existen en el archivo de excel.

max_col = ws_IVF.max_column

# Creamos una variable Count  y   Sum  que comiencen con valor cero

count = 0
suma = 0


# Iteramos de forma tal que para cada una de las "i" columnas, busque el valor
# que coincidan con la fila 8 (la cual contiene los Trimestre-Año), y lo
# agregue a la lista con nombre "serie_IVF_nombre"

for i in range(1, max_col + 1):
    cell_obj = ws_IVF.cell(row = 8, column = i)
    suma += count + 1
    #print(cell_obj.value)
    #print(suma)
    serie_IVF_nombre.append(cell_obj.value)


# Idem el bucle anterior pero ahora iterando con la fila que contiene NO los
# Trimestre-Año, sino que contiene los valores del IVF que corresponde.

for i in range(1, max_col + 1):
    cell_obj = ws_IVF.cell(row = 21, column = i)
    suma += count + 1
    #print(cell_obj.value)
    #print(suma)
    serie_IVF_valor.append(cell_obj.value)

# print(serie_IVF)


# Ahora le agregamos a la serie un valor específico en el index/posición 5,
# con el título de la serie.

serie_IVF_nombre.insert(5, "Trimestre-Año")
serie_IVF_valor.insert(5, "IVF del PIB")


# Ahora vamos a re-definir dos series como sub-conjuntos de las anteriores, de
# manera tal que nos muestre una parte de la información (desde el index = 5
# al final) y no toda.

serie_IVF_PIB_nombre = serie_IVF_nombre[5:]
serie_IVF_PIB_valor = serie_IVF_valor[5:]
print("La serie del IVF del PIB es:", serie_IVF_PIB_valor)


# Creamos una nueva solapa/hoja en el archivo de excel, donde luego
# guardaremos la información que hemos estado procesando a patir de las
# series originales.

ws_Seba = wb_obj['Seba']


# Ahora lo que queremos es que nos guarde en la hoja que hemos creado con el
# nombre "Seba", las series de Trimestre-Año y los correspondientes IVFs
# Es decir, queremos algo como:
# I Trim 2016      99.5
# II Trim 2016     101.5
# etc

# Para ello, crearemos un bucle en el que para cada "i" trimestre en el conjunto de
# valores (rango) entre cero y el largo de la lista serie_IVF_PIB_nombre (cuya
# extensión coincide con la de serie_IVF_PIB_valor), nos agregue a la hoja
# creada con nombre "Seba", una lista con dos listas dentro de ella, siendo
# la primer al Trimestre-Año y la segunda el IVF.

print("\n \n \n")

print("Exportando información a solapa 'Seba' del archivo excel.......")

for i in range(0, (len(serie_IVF_PIB_nombre))):
    ws_Seba.append([serie_IVF_PIB_nombre[i], serie_IVF_PIB_valor[i]])
    wb_obj.save('Actividades_IVF.xlsx')


print("\n \n \n")

print("Exportación procesada exitosamente")

print("\n \n \n")


#print(serie_IVF_PIB_valor)



print("Calculamos los promedios aritméticos para cada año")

print("\n \n \n")

año2016 = serie_IVF_PIB_valor[1:5]
#print(año2016)

prom2016 = sum(año2016)/len(año2016)
print(prom2016)




año2017 = serie_IVF_PIB_valor[5:9]
prom2017 = sum(año2017)/len(año2017)
print(prom2017)

año2018 = serie_IVF_PIB_valor[9:13]
prom2018 = sum(año2018)/len(año2018)
print(prom2018)


año2019 = serie_IVF_PIB_valor[13:17]
prom2019 = sum(año2019)/len(año2019)
print(prom2019)


año2020 = serie_IVF_PIB_valor[17:21]
prom2020 = sum(año2020)/len(año2020)
print(prom2020)


print("\n \n \n")

print("Calculamos los variación porcentual entre el año 2017 y 2016")

print("\n \n \n")

var17_16 = ((prom2017/prom2016)-1)*100
print("La variación del PIB Real 2017 VS 2016 fue de: ", "{0:.2f}".format(var17_16),"%")


print("\n \n \n")





serie_IVF_PIB_nombre_graph = serie_IVF_nombre[6:]
serie_IVF_PIB_valor_graph = serie_IVF_valor[6:]


media_movil = list()
media_movil_1 = list()

def moving_average(x, w):
    return np.convolve(x, np.ones(w), 'valid') / w


media_movil.append(moving_average(serie_IVF_PIB_valor_graph,4))


#print(media_movil[0])

media_movil_1 = media_movil[0]

#print(media_movil_1)


print(len(serie_IVF_PIB_valor_graph))
#print(len(media_movil_1))

media_movil_1 = np.append(media_movil_1, [99.67, 99.67, 99.67])

print(len(media_movil_1))

print(media_movil_1)





def fpp_2():

    x1 = serie_IVF_PIB_nombre_graph
    y1 = serie_IVF_PIB_valor_graph


    plt.plot(x1, y1, color='red', linestyle='dashdot', linewidth = 3,
             marker='*', markerfacecolor='yellow', markersize=12, label = "IVF original")


    # Creamos las coordenadas del segundo dibujo y le decimos que la llame
    # "FPP: mejora tec. en Vacas", ya que lo que se está representando
    # es un evento exógeno que genera mayor productividad en la producción
    # de vacas.

    x2 = serie_IVF_PIB_nombre_graph
    y2 = media_movil_1


    plt.plot(x2, y2, color='blue', linestyle='solid', linewidth = 3,
             marker='*', markerfacecolor='green', markersize=12,label = "IVF Promedios Móviles")


    # Le damos un nombre al eje de las abscisas ("x)" y ordenadas ("y")

    plt.xlabel('Trimestre-Año')

    plt.ylabel('IVF del PIB')


    # Le damos un noombre al gráfico

    plt.title('IVF trimestral y Movil Trimestral')


    # Mostramos el nombre de cada uno de los gráficos dibujados

    plt.legend()


    # Le damos la orden al programa que nos muestre el gráfico creado, con sus
    # respectivos ejes y título.

    plt.show()

fpp_2()


print("\n \n \n")



print("--------------************ FIN DEL CÓDIGO **********---------------------")
