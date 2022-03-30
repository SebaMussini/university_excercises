#---------------------************* FUENTES ***********--------------------


# https://www.geeksforgeeks.org/python-reading-excel-file-using-openpyxl-module/

# https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f




#------------------------------*********************------------------------

# Para leer un archivo de excel (xlsx), utilizaremos la librería
# openpyxl


# Importamos la librería

import openpyxl

from openpyxl import Workbook

from openpyxl import load_workbook




lista1 = list()


# Lo siguiente que haremos será indicarle al programa el lugar donde se
# encuentra el archivo de excel con el cual vamos a trabajar.

path = "/home/seba/Escritorio/Clases_UM_2022/Intro_Eco_2022/Práctico_1_Seba_Mussini/2/2_Python_RP_Smith/Smith.xlsx"



# Una vez realizado lo anterior, debemos crear un objeto en el cual
# que abra lo que existe en el path. A este objeto, lo llamaremos
#   "wb_obj"  // "WorkBook"

wb_obj = openpyxl.load_workbook(path)


# Pasamos ahora a crear un objeto que active las hojas que existen en el
#  WorkBook, al que llamaremos "sheet_obj"

sheet_obj = wb_obj.active


# Cell objects also have a row, column, and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.


# Ahora crearemos un objeto que nos traiga las celdas contenidas en la hoja
# que corresponda del archivo que corresponda, y lo hacemos a través del
# comando   sheet object's cell() method.

# Por ejemplo, queremos ver que contiene la celda de coordenadas
#  fila = 3    ,    columna = 2
# Nos debería mostrar el valor   "80"

cell_obj = sheet_obj.cell(row = 3, column = 2)

# Nos debería mostrar el valor   "80"

print("\n \n")

print("Mostrar el contenido de la celda con coordenadas Fila = 3 y Columna = 2 \n")

continuar = input("Para continuar, presione    'Enter'       ")

print("\n \n")

print(cell_obj.value)

print("\n \n")

# Podriamos querer conocer también las dimensiones de la matriz, es decir,
# el total de filas y columnas que ella contiene.

# Lo solicitamos a traves de los comandos "max.row" y "max.column"

print("Mostrar dimensiones de la matriz, filas y columnas \n")

continuar = input("Para continuar, presione    'Enter'       ")

print("\n \n")


print("Cantidad total de  filas: ", sheet_obj.max_row)

print("\n \n")

print("Cantidad total de columnas: ", sheet_obj.max_column)

print("\n \n")



# Podríamos también querer conocer los nombres de todas las columnas.

max_col = sheet_obj.max_column
max_row = sheet_obj.max_row

# Debemos crear un "loop finito" para iterar a lo largo de todas ellas.

# Se lee de la siguiente manera: para "i" en el rango de 1 al número máximo
# de columnas + 1, creamos un "objeto celdas" que nos traiga todas las
# columnas que existen a lo largo de la fila 1.
# Nota: si quisieramos un rango finito, por ejemplo, entre 1 y 5, se debe
# escribir como  "for i in range (1, 6)", porque la funcion range, se mueve
# entre el primer número y el inmediatamente anterior al señalado en el
# segundo argumento.

print("Los nombres de todas las columnas son:  \n")

continuar = input("Para continuar, presione    'Enter'       ")

print("\n \n")

for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row = 1, column = i)
    print(cell_obj.value)

print("\n \n")

print("Los nombres de todas las filas son:  \n")

continuar = input("Para continuar, presione    'Enter'       ")

print("\n \n")

for j in range(1, max_row + 1):
    cell_obj2 = sheet_obj.cell(row = j, column = 1)
    print(cell_obj2.value)

print("\n \n")


# Lo siguiente, es combinar lo visto en el paso anterior: veremos
# el contenido de toda la tabla. ¿Cómo lo haremos? Iterando a traves de
# todas las filas y todas las columnas.
# Para ello, deberemos crear un "loop finito" dentro de otro "loop finito"
#  (nested loop).

print("Mostrar todas las filas y columnas (toda la matriz) \n \n")

continuar = input("Para continuar, presione    'Enter'       ")

print("\n \n")

# Para cada "i" filas, en todas las filas
for i in range(1, max_row+1):
# Es decir, que por ejemplo, para al fila = 1, ejecuta lo que sigue

     # para cada "j" columnas, en todas las columnas
     for j in range(1, max_col+1):

          # queremos cada celda de coordenadas filas = i, columna = j
          cell_obj = sheet_obj.cell(row=i, column=j)
          # Entonces, obtendré las celdas:
          # fila = 1, columna = 1
          # fila = 1, columna = 2
          # fila = 1, columna = 3
          # fila = 1, columna = max_col + 1

          # al finalizar cada iteración, separarlas con ' | '
          print(cell_obj.value, end=' | ')




     # Al terminar con fila = 1 para todas las columnas, vuelve a comenzar
     # con fila = 2 y todas las columnas
     print('\n')



# Lo que haremos a continuación será "aislar" la cantidad de cada uno de
# los items de Sebastián y Federico, para así luego realizar cálculos
# dentro del ambiente Python.

print('\n \n')


for i in range(3, max_row+1):
# Es decir, que por ejemplo, para al fila = 1, ejecuta lo que sigue

     # para cada "j" columnas, en todas las columnas
     for j in range(1, max_col+1):

          # queremos cada celda de coordenadas filas = i, columna = j
          cell_obj = sheet_obj.cell(row=i, column=j)
          # Entonces, obtendré las celdas:
          # fila = 1, columna = 1
          # fila = 1, columna = 2
          # fila = 1, columna = 3
          # fila = 1, columna = max_col + 1

          # al finalizar cada iteración, separarlas con ' | '
          lista1.append(cell_obj.value)

     # print(lista1)

# En el paso anterior, el resultado es una gran lista de datos, que debemos
# separar en varias sub-listas para poder manipularlas.
# Observese que nos interesa un modelo de lista como el siguiente:
# ["Nombre", dotación de cocos, dotación de peces], por lo que separaremos
# la lista grande cada tres elementos.

n = 3

# range(0, len(test_list), n) devuelve un rango de números que comienzan en 0
# y terminan en len(test_list) con un paso de n. Por ejemplo, range(0, 10, 3)
# devolverá (0, 3, 6, 9)

lista2 = [lista1[i:i + n] for i in range(0, len(lista1), n)]


# Ahora separaré los elementos que me interesan de cada una de las sub-listas
# creadas.

continuar = input("Para continuar, presione    'Enter'       ")

print("\n \n")

sebastian = lista2[0]
dotseb = sebastian[1:]
# print(dotseb)
print("La dotación de Sebastián es de", dotseb[0], "cocos", "y", dotseb[1], "peces")

print('\n')

continuar = input("Para continuar, presione    'Enter'       ")

print("\n \n")

federico = lista2[1]
dotfed = federico[1:]
# print(dotfed)
print("La dotación de Federico es de", dotfed[0], "cocos", "y", dotfed[1], "peces")

print('\n \n')


continuar = input("Para continuar, presione    'Enter'       ")

print("\n \n")

print("Verifiquemos si existen ventajas absolutas.")

print("\n \n")

continuar = input("Para continuar, presione    'Enter'       ")

print("\n \n")

if dotseb[0] > dotfed[0] and dotseb[1] > dotfed[1]:
    print("Sebastián tiene ventajas absolutas en ambos bienes")
elif dotseb[0] > dotfed[0] or dotseb[1] > dotfed[1]:
    print("Sebastián tiene ventajas absolutas en UNO de los bienes")
else:
    print("Federico tiene ventajas absolutas en ambos bienes")


print('\n \n')

print("Calculemos los costos de oportunidad para ambos personajes.")

print("\n \n")

continuar = input("Para continuar, presione    'Enter'       ")

print("\n \n")

# Costo de oportunidad de un pescado para Sebastián.
# Es decir, "a cuantos cocos debe renunciar Sebastián para conseguir un
# pescado adicional"
# "cops" = "Costo de Oportunidad de un Pescado para Sebastián"


cops = (dotseb[0]/dotseb[1])
print("El costo de oportunidad de un pescado para Sebastián es de:", float(cops), "cocos")

print('\n')

cocs = (dotseb[1]/dotseb[0])
print("El costo de oportunidad de un coco para Sebastián es de:", float(cocs), "pescados")

print('\n')

copf = (dotfed[0]/dotfed[1])
print("El costo de oportunidad de un pescado para Federico es de:", float(copf), "cocos")

print('\n')

cocf = (dotfed[1]/dotfed[0])
print("El costo de oportunidad de un coco para Federico es de:", float(cocf), "pescados")

print('\n \n')



continuar = input("Para continuar, presione    'Enter'       ")

print("\n \n")


if (dotseb[0]/dotseb[1]) < (dotfed[0]/dotfed[1]):
    print("Seba tiene ventajas comparativas en pescar")
else:
    print("Federico tiene ventajas comparativas en pescar")


print("\n \n")


if (dotseb[1]/dotseb[0]) < (dotfed[1]/dotfed[0]):
    print("Seba tiene ventajas comparativas en juntar cocos")
else:
    print("Federico tiene ventajas comparativas en juntar cocos")

print("\n \n")

continuar = input("Para continuar, presione    'Enter'       ")


print('\n \n')

# Si dedican media jornada a cada uno de los productos mientras se encuentran
# en autarquía, ¿a cuánto ascenderan las respectivas dotaciones?

print("En Autarquía, dedicando la mitad de la jornada a cada uno de los bienes....")

print('\n \n')

continuar = input("Para continuar, presione    'Enter'       ")

print('\n \n')

jornada = 8

# PHS = Pescados por Hora Sebastián
# CHF = Cocos por Hora Federico

PHS = dotseb[1]/jornada
#print(PHS)

CHS = dotseb[0]/jornada
#print(CHS)

PHF = dotfed[1]/jornada
#print(PHF)

CHF = dotfed[0]/jornada
#print(CHF)

print("Sebastián tendría", PHS*4, "pescados y", CHS*4, "cocos")

print('\n \n')

print("Federico tendría", PHF*4, "pescados y", CHF*4, "cocos")

print('\n \n')

print("Si comerciaran y ambos produjeran solamente en aquello que tienen ventajas comparativas....")

print("Sebastián produciría", PHS*8, "pescados y", CHS*0, "cocos")

print('\n \n')

print("Federico produciría", PHF*0, "pescados y", CHF*8, "cocos")

print('\n \n')

# Sabemos que los valores de intercambio se encontrarán entre los costos
# de oportunidad respectivos en autarquía, por lo que Sebastián aceptará
# cambiar pescados por cocos siempre y cuando la tasa de cambio sea
# mayor a 10 cocos por 1 pescado, mientras Federico estará afin a este trueque
# siempre y cuando le cambien 1 pescado por menos de 15 cocos.
# Por tanto, habrá negociación y mutuo beneficio en un rango de cambio
#  10 + epsilon < 1 pescado < 15 - epsilon    ,  con epsilon ≈ 0
# A su vez, el inverso de la tasa de cambio para el pescado, será la fijada
# para los cocos, por existir una tasa de cambio constante entre ambas, pues es
# una función lineal.
# El único punto donde es igualmente beneficioso para ambos, es cuando se
# cambian 12,5 cocos por cada un pescado, lo que implica 0,08 pescados por cada
# un coco  (1/12,5)


continuar = input("Para continuar, presione    'Enter'       ")

print("\n \n")


print("Si la tasa de cambio pactada es de 12 cocos por cada 1 pescado, las dotaciones finales serán:")

print('\n \n')

continuar = input("Para continuar, presione    'Enter'       ")

print("\n \n")


print("Sebastián tendría", (PHS*8)-(4), "pescados y", (CHS*0)+(4*12), "cocos")

print('\n \n')

print("Federico tendría", (PHF*0)+(4), "pescados y", (CHF*8)-(12*4), "cocos")

print('\n \n')

print("************-------------------------FIN DEL CÓDIGO----------------**************")

print('\n \n')

salir = input("Para SALIR, presione    'Enter'       ")
